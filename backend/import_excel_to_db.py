from openpyxl import load_workbook
import psycopg

DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "dbname": "he_ho_tro_ra_quyet_dinh",
    "user": "postgres",
    "password": "Chihan0403",
}

EXCEL_FILE = r"..\supplier_ranking_grades.xlsx"

HEADER_TO_DB_NAME = {
    "chất lượng": "quality",
    "số lượng": "quantity",
    "điều kiện và phương thức thanh toán": "conditions_and_method_of_payment",
    "khả năng phục vụ và giao tiếp của nhà cung cấp": "serviceability_and_communicativeness_of_the_supplier",
    "uy tín và năng lực của nhà cung cấp": "reputation_of_the_supplier_and_its_competence",
    "tính linh hoạt": "flexibility",
    "tình hình tài chính của nhà cung cấp": "financial_condition_of_the_supplier",
    "tình trạng tài sản của nhà cung cấp": "condition_of_the_supplier_assets",
    "kết quả kinh doanh và số lượng nhân viên": "business_results_and_number_of_employees",
    "giá": "price",
    "thời gian giao hàng": "delivery_time",
    "vị trí và kết nối giao thông của nhà cung cấp": "supplier_location_and_traffic_connections",
}

SUPPLIER_DETAILS = {
    1: {
        "ten_ncc": "Công ty TNHH Vật tư Minh Phát",
        "dia_chi": "12 Nguyễn Trãi, Thanh Xuân, Hà Nội",
        "so_dien_thoai": "02438256801",
        "email": "lienhe@minhphat.vn",
        "mo_ta": "Nhà cung cấp vật tư tổng hợp, phù hợp đơn hàng số lượng vừa và nhỏ.",
    },
    2: {
        "ten_ncc": "Công ty Cổ phần Thương mại Hưng Thịnh",
        "dia_chi": "85 Điện Biên Phủ, Quận 3, TP. Hồ Chí Minh",
        "so_dien_thoai": "02839301202",
        "email": "sales@hungthinh.com.vn",
        "mo_ta": "Đơn vị thương mại có mạng lưới phân phối rộng, phản hồi báo giá nhanh.",
    },
    3: {
        "ten_ncc": "Công ty TNHH Kỹ thuật Đông Á",
        "dia_chi": "44 Lê Duẩn, Hải Châu, Đà Nẵng",
        "so_dien_thoai": "02363881303",
        "email": "contact@dongatech.vn",
        "mo_ta": "Chuyên cung cấp thiết bị kỹ thuật với đội ngũ hỗ trợ triển khai tại chỗ.",
    },
    4: {
        "ten_ncc": "Công ty Cổ phần Công nghiệp Việt Thành",
        "dia_chi": "118 Trần Phú, Ninh Kiều, Cần Thơ",
        "so_dien_thoai": "02923821404",
        "email": "kinhdoanh@vietthanh.vn",
        "mo_ta": "Có kinh nghiệm cung ứng cho doanh nghiệp sản xuất và kho vận miền Tây.",
    },
    5: {
        "ten_ncc": "Công ty TNHH Đầu tư Nam Việt",
        "dia_chi": "27 Lạch Tray, Ngô Quyền, Hải Phòng",
        "so_dien_thoai": "02253876505",
        "email": "info@namvietgroup.vn",
        "mo_ta": "Nhà cung cấp linh hoạt về phương thức thanh toán và lịch giao hàng.",
    },
    6: {
        "ten_ncc": "Công ty Cổ phần Dịch vụ Quốc tế An Bình",
        "dia_chi": "63 Hùng Vương, Hồng Bàng, Hải Phòng",
        "so_dien_thoai": "02253911606",
        "email": "support@anbinhglobal.vn",
        "mo_ta": "Mạnh về phối hợp giao nhận liên tỉnh và xử lý đơn hàng gấp.",
    },
    7: {
        "ten_ncc": "Công ty TNHH Thương mại Đại Phúc",
        "dia_chi": "141 Nguyễn Văn Cừ, Long Biên, Hà Nội",
        "so_dien_thoai": "02436544707",
        "email": "dvkh@daiphuc.vn",
        "mo_ta": "Đơn vị cung ứng ổn định, phù hợp các đơn hàng lặp lại theo tháng.",
    },
    8: {
        "ten_ncc": "Công ty Cổ phần Logistic Thành Nam",
        "dia_chi": "52 Phạm Văn Đồng, Bắc Từ Liêm, Hà Nội",
        "so_dien_thoai": "02437892008",
        "email": "hello@thanhnamlog.vn",
        "mo_ta": "Có lợi thế về kết nối giao thông và thời gian giao hàng khu vực phía Bắc.",
    },
    9: {
        "ten_ncc": "Công ty TNHH Công nghệ và Thiết bị Á Châu",
        "dia_chi": "210 Võ Văn Kiệt, Quận 1, TP. Hồ Chí Minh",
        "so_dien_thoai": "02839102109",
        "email": "sales@achau-tech.vn",
        "mo_ta": "Cung cấp thiết bị công nghệ, hồ sơ năng lực và chứng từ đầy đủ.",
    },
    10: {
        "ten_ncc": "Công ty Cổ phần Vật tư Hoàng Gia",
        "dia_chi": "74 Bạch Đằng, Hải Châu, Đà Nẵng",
        "so_dien_thoai": "02363820110",
        "email": "contact@hoanggia.vn",
        "mo_ta": "Chuyên cung ứng vật tư dự án, có khả năng đáp ứng đơn hàng tiến độ cao.",
    },
    11: {
        "ten_ncc": "Công ty TNHH Sản xuất và Thương mại Tân Long",
        "dia_chi": "95 Quang Trung, Vinh, Nghệ An",
        "so_dien_thoai": "02383877111",
        "email": "info@tanlong.vn",
        "mo_ta": "Nhà cung cấp sản xuất trực tiếp, giá cạnh tranh cho đơn hàng lớn.",
    },
    12: {
        "ten_ncc": "Công ty Cổ phần Giải pháp Nguồn cung Việt",
        "dia_chi": "128 Trường Chinh, Thanh Xuân, Hà Nội",
        "so_dien_thoai": "02435621812",
        "email": "procurement@nguoncungviet.vn",
        "mo_ta": "Có năng lực tổng hợp đơn hàng từ nhiều nhóm sản phẩm và nhiều tỉnh thành.",
    },
    13: {
        "ten_ncc": "Công ty TNHH Phát triển Thái Dương",
        "dia_chi": "18 Lý Tự Trọng, Ninh Kiều, Cần Thơ",
        "so_dien_thoai": "02923810213",
        "email": "thai.duong@thaiduong.vn",
        "mo_ta": "Dịch vụ chăm sóc khách hàng tốt, theo dõi đơn hàng sát tiến độ.",
    },
    14: {
        "ten_ncc": "Công ty Cổ phần Đầu tư và Cung ứng Đông Nam",
        "dia_chi": "77 Phan Chu Trinh, Hoàn Kiếm, Hà Nội",
        "so_dien_thoai": "02439245114",
        "email": "sales@dongnamsupply.vn",
        "mo_ta": "Phù hợp các dự án cần hồ sơ pháp lý, hợp đồng và nghiệm thu chặt chẽ.",
    },
    15: {
        "ten_ncc": "Công ty TNHH Vận tải và Kho vận Gia Hưng",
        "dia_chi": "33 Tố Hữu, Hà Đông, Hà Nội",
        "so_dien_thoai": "02433572015",
        "email": "ops@giahunglogistics.vn",
        "mo_ta": "Thế mạnh về lưu kho và giao hàng phân tán đến nhiều điểm nhận.",
    },
    16: {
        "ten_ncc": "Công ty Cổ phần TMDV Bách Khoa",
        "dia_chi": "156 Nguyễn Hữu Cảnh, Bình Thạnh, TP. Hồ Chí Minh",
        "so_dien_thoai": "02838998116",
        "email": "kinhdoanh@bachkhoa.vn",
        "mo_ta": "Đơn vị thương mại dịch vụ có đội ngũ kỹ thuật hỗ trợ sau bán hàng.",
    },
    17: {
        "ten_ncc": "Công ty TNHH Hợp Nhất Sài Gòn",
        "dia_chi": "229 Cộng Hòa, Tân Bình, TP. Hồ Chí Minh",
        "so_dien_thoai": "02838122717",
        "email": "contact@hopnhat.vn",
        "mo_ta": "Mạnh về xử lý đơn hàng nhanh, chấp nhận nhiều hình thức giao nhận.",
    },
    18: {
        "ten_ncc": "Công ty Cổ phần Vina Source",
        "dia_chi": "68 Trần Hưng Đạo, Hải Châu, Đà Nẵng",
        "so_dien_thoai": "02363598118",
        "email": "hello@vinasource.vn",
        "mo_ta": "Nguồn hàng ổn định, có khả năng phối hợp với doanh nghiệp xuất nhập khẩu.",
    },
    19: {
        "ten_ncc": "Công ty TNHH Cung ứng Toàn Phát",
        "dia_chi": "91 Lê Hồng Phong, Ngô Quyền, Hải Phòng",
        "so_dien_thoai": "02253987219",
        "email": "sales@toanphat.vn",
        "mo_ta": "Nhà cung cấp thiên về giá tốt, thích hợp đơn hàng tiết kiệm ngân sách.",
    },
    20: {
        "ten_ncc": "Công ty Cổ phần Chuỗi Cung ứng An Tâm",
        "dia_chi": "104 Nguyễn Văn Linh, Thanh Khê, Đà Nẵng",
        "so_dien_thoai": "02363726420",
        "email": "info@antamsupply.vn",
        "mo_ta": "Cung cấp trọn gói từ đặt hàng đến giao nhận và đối soát chứng từ.",
    },
    21: {
        "ten_ncc": "Công ty TNHH Thiết bị Hòa Bình",
        "dia_chi": "40 Xô Viết Nghệ Tĩnh, Bình Thạnh, TP. Hồ Chí Minh",
        "so_dien_thoai": "02838456221",
        "email": "care@hoabinh-equip.vn",
        "mo_ta": "Có danh mục thiết bị đa dạng, phù hợp mua sắm theo từng giai đoạn.",
    },
    22: {
        "ten_ncc": "Công ty Cổ phần Tân Tiến Logistics",
        "dia_chi": "15 Nguyễn Văn Linh, Lê Chân, Hải Phòng",
        "so_dien_thoai": "02253631122",
        "email": "service@tantienlog.vn",
        "mo_ta": "Mạnh về hậu cần và tối ưu lộ trình giao hàng khu vực cảng.",
    },
    23: {
        "ten_ncc": "Công ty TNHH Dịch vụ Công nghiệp Phương Nam",
        "dia_chi": "166 Phạm Hùng, Nam Từ Liêm, Hà Nội",
        "so_dien_thoai": "02437668223",
        "email": "support@phuongnamind.vn",
        "mo_ta": "Chuyên phục vụ doanh nghiệp công nghiệp với quy trình giao nhận tiêu chuẩn.",
    },
    24: {
        "ten_ncc": "Công ty Cổ phần Giải pháp Chuỗi giá trị Việt",
        "dia_chi": "57 Nguyễn Thị Minh Khai, Quận 1, TP. Hồ Chí Minh",
        "so_dien_thoai": "02839333124",
        "email": "contact@valuechainviet.vn",
        "mo_ta": "Có năng lực tư vấn và đồng hành cùng bộ phận mua hàng doanh nghiệp.",
    },
    25: {
        "ten_ncc": "Công ty TNHH Kinh doanh Tổng hợp Hoàng Minh",
        "dia_chi": "21 Lê Lợi, TP. Huế",
        "so_dien_thoai": "02343892125",
        "email": "info@hoangminh.vn",
        "mo_ta": "Nhà cung cấp tổng hợp, thuận tiện cho đơn hàng khu vực miền Trung.",
    },
    26: {
        "ten_ncc": "Công ty Cổ phần Vật tư và Thiết bị Đại Nam",
        "dia_chi": "73 Hàm Nghi, Thanh Khê, Đà Nẵng",
        "so_dien_thoai": "02363781526",
        "email": "sales@dainam.vn",
        "mo_ta": "Giá tốt, danh mục sản phẩm đa dạng, phù hợp đấu thầu mua sắm cơ bản.",
    },
    27: {
        "ten_ncc": "Công ty TNHH Sản phẩm Công nghiệp Bắc Nam",
        "dia_chi": "188 Nguyễn Oanh, Gò Vấp, TP. Hồ Chí Minh",
        "so_dien_thoai": "02839845027",
        "email": "hotro@bacnam.vn",
        "mo_ta": "Có năng lực sản xuất bổ sung theo yêu cầu riêng của khách hàng.",
    },
    28: {
        "ten_ncc": "Công ty Cổ phần Hợp tác Nguồn lực Việt",
        "dia_chi": "132 Hoàng Quốc Việt, Cầu Giấy, Hà Nội",
        "so_dien_thoai": "02437561228",
        "email": "business@nguonlucviet.vn",
        "mo_ta": "Nhà cung cấp có quy trình xử lý chứng từ và thanh toán tương đối linh hoạt.",
    },
    29: {
        "ten_ncc": "Công ty TNHH Thương mại Kết Nối Mới",
        "dia_chi": "49 Trần Não, TP. Thủ Đức, TP. Hồ Chí Minh",
        "so_dien_thoai": "02837445529",
        "email": "sales@ketnoimoi.vn",
        "mo_ta": "Phù hợp các doanh nghiệp cần nguồn hàng thay thế và phương án dự phòng.",
    },
    30: {
        "ten_ncc": "Công ty Cổ phần Chuẩn Việt Supply",
        "dia_chi": "24 Võ Nguyên Giáp, Sơn Trà, Đà Nẵng",
        "so_dien_thoai": "02363974230",
        "email": "info@chuanviet.vn",
        "mo_ta": "Lợi thế về vị trí kho và giao nhận nhanh cho khu vực miền Trung.",
    },
    31: {
        "ten_ncc": "Công ty TNHH Thương mại - Dịch vụ Bảo Tín",
        "dia_chi": "109 Lê Quang Đạo, Nam Từ Liêm, Hà Nội",
        "so_dien_thoai": "02432011831",
        "email": "baotin@baotin.vn",
        "mo_ta": "Dịch vụ hậu mãi tốt, phù hợp đơn hàng cần theo dõi sau giao nhận.",
    },
    32: {
        "ten_ncc": "Công ty Cổ phần Công nghiệp Sông Hồng",
        "dia_chi": "55 Trần Thái Tông, Cầu Giấy, Hà Nội",
        "so_dien_thoai": "02437992132",
        "email": "contact@songhongind.vn",
        "mo_ta": "Doanh nghiệp công nghiệp có năng lực cung ứng đều đặn theo hợp đồng khung.",
    },
    33: {
        "ten_ncc": "Công ty TNHH Hạ tầng và Cung ứng Miền Nam",
        "dia_chi": "87 Nguyễn Văn Linh, Quận 7, TP. Hồ Chí Minh",
        "so_dien_thoai": "02837762133",
        "email": "service@miennamcorp.vn",
        "mo_ta": "Mạnh về phục vụ doanh nghiệp khu công nghiệp phía Nam.",
    },
    34: {
        "ten_ncc": "Công ty Cổ phần Thương mại Duy Tân",
        "dia_chi": "36 Lê Duẩn, TP. Vinh, Nghệ An",
        "so_dien_thoai": "02383811234",
        "email": "duytan@duytan.vn",
        "mo_ta": "Có mạng lưới đối tác vùng Bắc Trung Bộ, phù hợp giao hàng nhiều điểm.",
    },
    35: {
        "ten_ncc": "Công ty TNHH Nguồn Hàng Việt Phú",
        "dia_chi": "145 Quốc lộ 13, Thuận An, Bình Dương",
        "so_dien_thoai": "02743654535",
        "email": "kinhdoanh@vietphu.vn",
        "mo_ta": "Nhà cung cấp có kho gần khu công nghiệp, thuận tiện đơn hàng gấp và bổ sung.",
    },
}


def update_supplier_profiles(cur):
    for supplier_id, detail in SUPPLIER_DETAILS.items():
        cur.execute(
            """
            UPDATE public.nha_cung_cap
            SET ten_ncc = %s,
                dia_chi = %s,
                so_dien_thoai = %s,
                email = %s,
                mo_ta = %s
            WHERE ma_ncc = %s
            """,
            (
                detail["ten_ncc"],
                detail["dia_chi"],
                detail["so_dien_thoai"],
                detail["email"],
                detail["mo_ta"],
                supplier_id,
            ),
        )


def main():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("File Excel rỗng.")
        return

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    print("Sheet đang đọc:", wb.sheetnames[0])
    print("Headers:", headers)
    print("Số dòng dữ liệu:", len(rows) - 1)

    with psycopg.connect(
        host=DB_CONFIG["host"],
        port=DB_CONFIG["port"],
        dbname=DB_CONFIG["dbname"],
        user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
    ) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT ma_tieu_chi, ten_tieu_chi FROM public.tieu_chi ORDER BY ma_tieu_chi;"
            )
            criteria_rows = cur.fetchall()
            criteria_map = {name: ma for ma, name in criteria_rows}

            print("Tiêu chí trong DB:", criteria_map)

            update_supplier_profiles(cur)
            cur.execute("DELETE FROM public.danh_gia_ncc;")

            inserted = 0

            for row in rows[1:]:
                if row[0] is None:
                    continue

                supplier_id = int(row[0])

                for col_idx in range(1, len(headers)):
                    header = headers[col_idx]
                    mapped_name = HEADER_TO_DB_NAME.get(header)

                    if not mapped_name:
                        continue

                    if mapped_name not in criteria_map:
                        print(f"Không tìm thấy tiêu chí trong DB: {mapped_name}")
                        continue

                    diem = row[col_idx]
                    if diem is None:
                        continue

                    ma_tieu_chi = criteria_map[mapped_name]
                    cur.execute(
                        """
                        INSERT INTO public.danh_gia_ncc (ma_ncc, ma_tieu_chi, diem)
                        VALUES (%s, %s, %s)
                        """,
                        (supplier_id, ma_tieu_chi, float(diem)),
                    )
                    inserted += 1

        conn.commit()

    print(f"Đã cập nhật hồ sơ nhà cung cấp và import {inserted} dòng vào bảng danh_gia_ncc.")


if __name__ == "__main__":
    main()
